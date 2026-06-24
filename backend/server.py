from fastapi import FastAPI, APIRouter, HTTPException, Depends, Header, Request
from fastapi.responses import StreamingResponse, PlainTextResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import io
import csv
import logging
import urllib.parse
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict
import uuid
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

# ----------------- Constants -----------------
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "admin123"
ADMIN_TOKEN = "iminationz-admin-token-2026"
STORE_NAME = "Iminationz"
OWNER_WHATSAPP_NUMBERS = ["9044625875", "8188996721"]

IST = timezone(timedelta(hours=5, minutes=30))


def now_ist() -> datetime:
    return datetime.now(IST)


def to_dt_parts(dt: datetime):
    return {
        "date": dt.strftime("%Y-%m-%d"),
        "day": dt.strftime("%A"),
        "time": dt.strftime("%H:%M:%S"),
        "iso": dt.isoformat(),
    }


# ----------------- Auth -----------------
class LoginPayload(BaseModel):
    username: str
    password: str


def require_auth(request: Request, authorization: Optional[str] = Header(None)):
    token = None
    if authorization:
        token = authorization.replace("Bearer ", "").strip()
    if not token:
        token = request.query_params.get("_t")
    if not token:
        raise HTTPException(status_code=401, detail="Missing token")
    if token != ADMIN_TOKEN:
        raise HTTPException(status_code=401, detail="Invalid token")
    return True


@api_router.post("/auth/login")
async def login(payload: LoginPayload):
    if payload.username == ADMIN_USERNAME and payload.password == ADMIN_PASSWORD:
        return {"token": ADMIN_TOKEN, "store_name": STORE_NAME}
    raise HTTPException(status_code=401, detail="Invalid credentials")


# ----------------- Inventory -----------------
class InventoryCreate(BaseModel):
    item_id: str
    category: str
    item_name: str
    price: float
    cost_price: Optional[float] = 0
    opening_qty: int
    current_qty: Optional[int] = None


class InventoryUpdate(BaseModel):
    category: Optional[str] = None
    item_name: Optional[str] = None
    price: Optional[float] = None
    cost_price: Optional[float] = None
    opening_qty: Optional[int] = None
    current_qty: Optional[int] = None


class InventoryItem(BaseModel):
    id: str
    item_id: str
    category: str
    item_name: str
    price: float
    cost_price: float
    opening_qty: int
    current_qty: int
    sold_qty: int
    created_date: str
    last_updated: str


def inv_doc_to_model(doc: dict) -> dict:
    return {
        "id": doc["id"],
        "item_id": doc["item_id"],
        "category": doc["category"],
        "item_name": doc["item_name"],
        "price": doc["price"],
        "cost_price": float(doc.get("cost_price") or 0),
        "opening_qty": doc["opening_qty"],
        "current_qty": doc["current_qty"],
        "sold_qty": doc["sold_qty"],
        "created_date": doc["created_date"],
        "last_updated": doc["last_updated"],
    }


@api_router.get("/inventory", response_model=List[InventoryItem])
async def list_inventory(_: bool = Depends(require_auth)):
    docs = await db.inventory.find({}, {"_id": 0}).to_list(2000)
    docs.sort(key=lambda d: (d.get("category", ""), d.get("item_name", "")))
    return [inv_doc_to_model(d) for d in docs]


@api_router.post("/inventory", response_model=InventoryItem)
async def create_inventory(payload: InventoryCreate, _: bool = Depends(require_auth)):
    normalized_item_id = payload.item_id.strip().upper()
    existing = await db.inventory.find_one({"item_id": normalized_item_id}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Item ID already exists")
    now = now_ist()
    current_qty = payload.current_qty if payload.current_qty is not None else payload.opening_qty
    doc = {
        "id": str(uuid.uuid4()),
        "item_id": normalized_item_id,
        "category": payload.category.strip(),
        "item_name": payload.item_name.strip(),
        "price": float(payload.price),
        "cost_price": float(payload.cost_price or 0),
        "opening_qty": int(payload.opening_qty),
        "current_qty": int(current_qty),
        "sold_qty": 0,
        "created_date": now.isoformat(),
        "last_updated": now.isoformat(),
    }
    await db.inventory.insert_one(dict(doc))
    return inv_doc_to_model(doc)


@api_router.put("/inventory/{item_id}", response_model=InventoryItem)
async def update_inventory(item_id: str, payload: InventoryUpdate, _: bool = Depends(require_auth)):
    existing = await db.inventory.find_one({"id": item_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Item not found")
    update = {k: v for k, v in payload.dict(exclude_unset=True).items() if v is not None}
    if "opening_qty" in update:
        update["opening_qty"] = int(update["opening_qty"])
    if "current_qty" in update:
        update["current_qty"] = int(update["current_qty"])
    if "price" in update:
        update["price"] = float(update["price"])
    if "cost_price" in update:
        update["cost_price"] = float(update["cost_price"])
    update["last_updated"] = now_ist().isoformat()
    await db.inventory.update_one({"id": item_id}, {"$set": update})
    doc = await db.inventory.find_one({"id": item_id}, {"_id": 0})
    return inv_doc_to_model(doc)


@api_router.delete("/inventory/{item_id}")
async def delete_inventory(item_id: str, _: bool = Depends(require_auth)):
    res = await db.inventory.delete_one({"id": item_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    return {"deleted": True}


# ----------------- Billing -----------------
class BillItem(BaseModel):
    inv_id: str
    item_id: str
    item_name: str
    price: float
    qty: int
    line_total: float


class BillCreate(BaseModel):
    customer_mobile: Optional[str] = None
    customer_name: Optional[str] = None
    items: List[BillItem]
    cash_amount: float
    upi_amount: float


class Bill(BaseModel):
    id: str
    bill_number: str
    customer_mobile: Optional[str]
    customer_name: Optional[str] = None
    date: str
    day: str
    time: str
    iso: str
    items: List[BillItem]
    gross_amount: float
    discount: float
    final_amount: float
    cash_amount: float
    upi_amount: float
    payment_status: str


def calc_amounts(items: List[BillItem]):
    gross = round(sum(i.price * i.qty for i in items), 2)
    discount = round(gross * 0.10, 2) if gross > 699 else 0.0
    final_amount = round(gross - discount, 2)
    return gross, discount, final_amount


async def next_bill_number(date_str: str) -> str:
    compact = date_str.replace("-", "")
    prefix = f"BILL-{compact}-"
    count = await db.bills.count_documents({"bill_number": {"$regex": f"^{prefix}"}})
    return f"{prefix}{str(count + 1).zfill(3)}"


@api_router.post("/bills", response_model=Bill)
async def create_bill(payload: BillCreate, _: bool = Depends(require_auth)):
    if not payload.items or len(payload.items) == 0:
        raise HTTPException(status_code=400, detail="Bill must have at least one item")

    # Validate stock & compute line totals
    normalized: List[BillItem] = []
    for it in payload.items:
        if it.qty <= 0:
            raise HTTPException(status_code=400, detail=f"Invalid qty for {it.item_name}")
        inv = await db.inventory.find_one({"id": it.inv_id}, {"_id": 0})
        if not inv:
            raise HTTPException(status_code=400, detail=f"Item {it.item_name} not found in inventory")
        if inv["current_qty"] < it.qty:
            raise HTTPException(
                status_code=400,
                detail=f"Insufficient stock for {inv['item_name']} (available {inv['current_qty']})",
            )
        line_total = round(inv["price"] * it.qty, 2)
        normalized.append(BillItem(
            inv_id=inv["id"],
            item_id=inv["item_id"],
            item_name=inv["item_name"],
            price=float(inv["price"]),
            qty=int(it.qty),
            line_total=line_total,
        ))

    gross, discount, final_amount = calc_amounts(normalized)
    paid_total = round(payload.cash_amount + payload.upi_amount, 2)
    if abs(paid_total - final_amount) > 0.01:
        raise HTTPException(
            status_code=400,
            detail=f"Cash + UPI ({paid_total}) must equal Final Amount ({final_amount})",
        )

    now = now_ist()
    parts = to_dt_parts(now)
    bill_number = await next_bill_number(parts["date"])

    # Atomically deduct inventory (sequential per item using $inc with conditional check)
    deducted = []
    for it in normalized:
        res = await db.inventory.update_one(
            {"id": it.inv_id, "current_qty": {"$gte": it.qty}},
            {
                "$inc": {"current_qty": -it.qty, "sold_qty": it.qty},
                "$set": {"last_updated": now.isoformat()},
            },
        )
        if res.modified_count == 0:
            # Rollback previously deducted items
            for d in deducted:
                await db.inventory.update_one(
                    {"id": d["inv_id"]},
                    {"$inc": {"current_qty": d["qty"], "sold_qty": -d["qty"]}},
                )
            raise HTTPException(status_code=400, detail=f"Stock changed for {it.item_name}, please retry")
        deducted.append({"inv_id": it.inv_id, "qty": it.qty})

    bill_doc = {
        "id": str(uuid.uuid4()),
        "bill_number": bill_number,
        "customer_mobile": payload.customer_mobile,
        "customer_name": (payload.customer_name or "").strip() or None,
        "date": parts["date"],
        "day": parts["day"],
        "time": parts["time"],
        "iso": parts["iso"],
        "items": [i.dict() for i in normalized],
        "gross_amount": gross,
        "discount": discount,
        "final_amount": final_amount,
        "cash_amount": round(payload.cash_amount, 2),
        "upi_amount": round(payload.upi_amount, 2),
        "payment_status": "PAID",
    }
    await db.bills.insert_one(dict(bill_doc))
    bill_doc.pop("_id", None)
    return Bill(**bill_doc)


@api_router.get("/bills", response_model=List[Bill])
async def list_bills(
    filter: Optional[str] = "today",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    search: Optional[str] = None,
    _: bool = Depends(require_auth),
):
    today = now_ist().strftime("%Y-%m-%d")
    yesterday = (now_ist() - timedelta(days=1)).strftime("%Y-%m-%d")
    month_start = now_ist().strftime("%Y-%m-01")

    query: dict = {}
    if filter == "today":
        query["date"] = today
    elif filter == "yesterday":
        query["date"] = yesterday
    elif filter == "month":
        query["date"] = {"$gte": month_start, "$lte": today}
    elif filter == "custom" and start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    elif filter == "all":
        pass

    if search:
        s = search.strip()
        query["$or"] = [
            {"bill_number": {"$regex": s, "$options": "i"}},
            {"customer_mobile": {"$regex": s, "$options": "i"}},
        ]

    docs = await db.bills.find(query, {"_id": 0}).sort("iso", -1).to_list(2000)
    return [Bill(**d) for d in docs]


@api_router.get("/bills/{bill_id}", response_model=Bill)
async def get_bill(bill_id: str, _: bool = Depends(require_auth)):
    doc = await db.bills.find_one({"id": bill_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Bill not found")
    return Bill(**doc)


# ----------------- Dashboard -----------------
@api_router.get("/dashboard/today")
async def dashboard_today(_: bool = Depends(require_auth)):
    today = now_ist().strftime("%Y-%m-%d")
    bills = await db.bills.find({"date": today}, {"_id": 0}).to_list(5000)
    total_sales = round(sum(b["final_amount"] for b in bills), 2)
    total_cash = round(sum(b["cash_amount"] for b in bills), 2)
    total_upi = round(sum(b["upi_amount"] for b in bills), 2)
    total_bills = len(bills)
    items_sold = sum(sum(i["qty"] for i in b["items"]) for b in bills)
    discount_given = round(sum(b["discount"] for b in bills), 2)
    avg_bill = round(total_sales / total_bills, 2) if total_bills else 0

    inv = await db.inventory.find({}, {"_id": 0}).to_list(5000)
    total_inventory_qty = sum(i["current_qty"] for i in inv)
    low_stock_count = sum(1 for i in inv if i["current_qty"] <= 5)

    return {
        "date": today,
        "total_sales": total_sales,
        "total_cash": total_cash,
        "total_upi": total_upi,
        "total_bills": total_bills,
        "items_sold": items_sold,
        "discount_given": discount_given,
        "average_bill_value": avg_bill,
        "total_inventory_qty": total_inventory_qty,
        "low_stock_count": low_stock_count,
        "store_name": STORE_NAME,
    }


# ----------------- Reports -----------------
@api_router.get("/reports/daily")
async def daily_report(date: Optional[str] = None, _: bool = Depends(require_auth)):
    target = date or now_ist().strftime("%Y-%m-%d")
    bills = await db.bills.find({"date": target}, {"_id": 0}).to_list(5000)
    total_sales = round(sum(b["final_amount"] for b in bills), 2)
    total_cash = round(sum(b["cash_amount"] for b in bills), 2)
    total_upi = round(sum(b["upi_amount"] for b in bills), 2)
    total_bills = len(bills)
    items_sold = sum(sum(i["qty"] for i in b["items"]) for b in bills)
    discount_given = round(sum(b["discount"] for b in bills), 2)
    avg_bill = round(total_sales / total_bills, 2) if total_bills else 0
    return {
        "date": target,
        "total_bills": total_bills,
        "total_sales": total_sales,
        "total_cash": total_cash,
        "total_upi": total_upi,
        "discount_given": discount_given,
        "items_sold": items_sold,
        "average_bill_value": avg_bill,
    }


@api_router.get("/reports/inventory")
async def inventory_report(_: bool = Depends(require_auth)):
    docs = await db.inventory.find({}, {"_id": 0}).to_list(5000)
    docs.sort(key=lambda d: (d.get("category", ""), d.get("item_name", "")))
    total_opening = sum(d["opening_qty"] for d in docs)
    total_current = sum(d["current_qty"] for d in docs)
    total_sold = sum(d["sold_qty"] for d in docs)
    low_stock = [inv_doc_to_model(d) for d in docs if d["current_qty"] <= 5]
    items = [inv_doc_to_model(d) for d in docs]
    return {
        "items": items,
        "summary": {
            "total_opening": total_opening,
            "total_current": total_current,
            "total_sold": total_sold,
            "low_stock_count": len(low_stock),
        },
        "low_stock": low_stock,
    }


# ----------------- Customer Lookup -----------------
@api_router.get("/customers/{mobile}")
async def customer_lookup(mobile: str, _: bool = Depends(require_auth)):
    """Return aggregate info for a customer mobile number — used to surface repeat customers."""
    bills = await db.bills.find({"customer_mobile": mobile}, {"_id": 0}).sort("iso", -1).to_list(2000)
    if not bills:
        return {
            "mobile": mobile,
            "is_returning": False,
            "visits": 0,
            "total_spent": 0,
            "last_visit": None,
            "last_name": None,
        }
    total_spent = round(sum(b["final_amount"] for b in bills), 2)
    last_name = next((b.get("customer_name") for b in bills if b.get("customer_name")), None)
    return {
        "mobile": mobile,
        "is_returning": True,
        "visits": len(bills),
        "total_spent": total_spent,
        "last_visit": bills[0]["date"],
        "last_name": last_name,
    }


# ----------------- Category Analytics -----------------
@api_router.get("/reports/category")
async def category_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    _: bool = Depends(require_auth),
):
    query: dict = {}
    if start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    bills = await db.bills.find(query, {"_id": 0}).to_list(10000)
    inv = await db.inventory.find({}, {"_id": 0}).to_list(5000)
    cost_by_inv_id: Dict[str, float] = {i["item_id"]: float(i.get("cost_price") or 0) for i in inv}
    cat_by_inv_id: Dict[str, str] = {i["item_id"]: i["category"] for i in inv}

    agg: Dict[str, Dict[str, float]] = {}
    for b in bills:
        for it in b["items"]:
            cat = cat_by_inv_id.get(it["item_id"], "Unknown")
            row = agg.setdefault(cat, {"qty": 0, "revenue": 0.0, "cost": 0.0})
            row["qty"] += it["qty"]
            row["revenue"] += it["line_total"]
            row["cost"] += cost_by_inv_id.get(it["item_id"], 0) * it["qty"]

    result = []
    for cat, row in agg.items():
        revenue = round(row["revenue"], 2)
        cost = round(row["cost"], 2)
        profit = round(revenue - cost, 2)
        margin = round((profit / revenue * 100), 2) if revenue > 0 else 0
        result.append({
            "category": cat,
            "qty_sold": int(row["qty"]),
            "revenue": revenue,
            "cost": cost,
            "profit": profit,
            "margin_pct": margin,
        })
    result.sort(key=lambda x: x["revenue"], reverse=True)
    return {"rows": result}


# ----------------- Exports -----------------
async def _filtered_bills(filter_: str, start_date: Optional[str], end_date: Optional[str]):
    today = now_ist().strftime("%Y-%m-%d")
    yesterday = (now_ist() - timedelta(days=1)).strftime("%Y-%m-%d")
    month_start = now_ist().strftime("%Y-%m-01")
    query: dict = {}
    if filter_ == "today":
        query["date"] = today
    elif filter_ == "yesterday":
        query["date"] = yesterday
    elif filter_ == "month":
        query["date"] = {"$gte": month_start, "$lte": today}
    elif filter_ == "custom" and start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    return await db.bills.find(query, {"_id": 0}).sort("iso", -1).to_list(20000)


def _set_header_style(ws, cols):
    fill = PatternFill("solid", fgColor="D4AF37")
    font = Font(bold=True, color="1A1500")
    for idx, _label in enumerate(cols, 1):
        cell = ws.cell(row=1, column=idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


@api_router.get("/exports/sales.xlsx")
async def export_sales_xlsx(
    filter: str = "month",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    _: bool = Depends(require_auth),
):
    bills = await _filtered_bills(filter, start_date, end_date)
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Sales"
    headers = ["Bill No", "Date", "Day", "Time", "Customer Name", "Mobile", "Items", "Gross", "Discount", "Final", "Cash", "UPI", "Status"]
    ws1.append(headers)
    _set_header_style(ws1, headers)
    for b in bills:
        items_str = "; ".join(f"{i['item_name']} x{i['qty']}" for i in b["items"])
        ws1.append([
            b["bill_number"], b["date"], b["day"], b["time"],
            b.get("customer_name") or "", b.get("customer_mobile") or "",
            items_str, b["gross_amount"], b["discount"], b["final_amount"],
            b["cash_amount"], b["upi_amount"], b["payment_status"],
        ])

    ws2 = wb.create_sheet("Line Items")
    line_headers = ["Bill No", "Date", "Item ID", "Item Name", "Qty", "Price", "Line Total"]
    ws2.append(line_headers)
    _set_header_style(ws2, line_headers)
    for b in bills:
        for it in b["items"]:
            ws2.append([
                b["bill_number"], b["date"], it["item_id"], it["item_name"],
                it["qty"], it["price"], it["line_total"],
            ])

    ws3 = wb.create_sheet("Summary")
    total_sales = round(sum(b["final_amount"] for b in bills), 2)
    total_cash = round(sum(b["cash_amount"] for b in bills), 2)
    total_upi = round(sum(b["upi_amount"] for b in bills), 2)
    total_discount = round(sum(b["discount"] for b in bills), 2)
    items_sold = sum(sum(i["qty"] for i in b["items"]) for b in bills)
    summary_headers = ["Metric", "Value"]
    ws3.append(summary_headers)
    _set_header_style(ws3, summary_headers)
    ws3.append(["Filter", filter])
    ws3.append(["Total Bills", len(bills)])
    ws3.append(["Total Sales", total_sales])
    ws3.append(["Cash", total_cash])
    ws3.append(["UPI", total_upi])
    ws3.append(["Discount Given", total_discount])
    ws3.append(["Items Sold", items_sold])

    for ws in (ws1, ws2, ws3):
        for col in ws.columns:
            length = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, 10), 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"iminationz_sales_{filter}_{now_ist().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@api_router.get("/exports/sales.csv")
async def export_sales_csv(
    filter: str = "month",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    _: bool = Depends(require_auth),
):
    bills = await _filtered_bills(filter, start_date, end_date)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Bill No", "Date", "Day", "Time", "Customer Name", "Mobile", "Items", "Gross", "Discount", "Final", "Cash", "UPI", "Status"])
    for b in bills:
        items_str = "; ".join(f"{i['item_name']} x{i['qty']}" for i in b["items"])
        w.writerow([
            b["bill_number"], b["date"], b["day"], b["time"],
            b.get("customer_name") or "", b.get("customer_mobile") or "",
            items_str, b["gross_amount"], b["discount"], b["final_amount"],
            b["cash_amount"], b["upi_amount"], b["payment_status"],
        ])
    fname = f"iminationz_sales_{filter}_{now_ist().strftime('%Y%m%d_%H%M%S')}.csv"
    return PlainTextResponse(
        buf.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@api_router.get("/exports/inventory.xlsx")
async def export_inventory_xlsx(_: bool = Depends(require_auth)):
    docs = await db.inventory.find({}, {"_id": 0}).to_list(5000)
    docs.sort(key=lambda d: (d.get("category", ""), d.get("item_name", "")))
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    headers = ["Item ID", "Category", "Name", "Price", "Cost Price", "Opening Qty", "Current Qty", "Sold Qty", "Low Stock", "Created", "Updated"]
    ws.append(headers)
    _set_header_style(ws, headers)
    for d in docs:
        ws.append([
            d["item_id"], d["category"], d["item_name"], d["price"], float(d.get("cost_price") or 0),
            d["opening_qty"], d["current_qty"], d["sold_qty"],
            "YES" if d["current_qty"] <= 5 else "",
            d["created_date"], d["last_updated"],
        ])
    for col in ws.columns:
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, 10), 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"iminationz_inventory_{now_ist().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@api_router.get("/exports/inventory.csv")
async def export_inventory_csv(_: bool = Depends(require_auth)):
    docs = await db.inventory.find({}, {"_id": 0}).to_list(5000)
    docs.sort(key=lambda d: (d.get("category", ""), d.get("item_name", "")))
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Item ID", "Category", "Name", "Price", "Cost Price", "Opening Qty", "Current Qty", "Sold Qty", "Low Stock"])
    for d in docs:
        w.writerow([
            d["item_id"], d["category"], d["item_name"], d["price"], float(d.get("cost_price") or 0),
            d["opening_qty"], d["current_qty"], d["sold_qty"],
            "YES" if d["current_qty"] <= 5 else "",
        ])
    fname = f"iminationz_inventory_{now_ist().strftime('%Y%m%d_%H%M%S')}.csv"
    return PlainTextResponse(
        buf.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ----------------- WhatsApp closing summary -----------------
@api_router.get("/whatsapp/closing")
async def whatsapp_closing(date: Optional[str] = None, _: bool = Depends(require_auth)):
    target = date or now_ist().strftime("%Y-%m-%d")
    bills = await db.bills.find({"date": target}, {"_id": 0}).to_list(5000)
    total_sales = round(sum(b["final_amount"] for b in bills), 2)
    total_cash = round(sum(b["cash_amount"] for b in bills), 2)
    total_upi = round(sum(b["upi_amount"] for b in bills), 2)
    items_sold = sum(sum(i["qty"] for i in b["items"]) for b in bills)
    discount = round(sum(b["discount"] for b in bills), 2)

    lines = [
        f"*{STORE_NAME} — Daily Closing*",
        f"Date: {target}",
        "",
        f"Bills: {len(bills)}",
        f"Total Sales: ₹{total_sales:,.2f}",
        f"Cash: ₹{total_cash:,.2f}",
        f"UPI: ₹{total_upi:,.2f}",
        f"Discount Given: ₹{discount:,.2f}",
        f"Items Sold: {items_sold}",
    ]
    message = "\n".join(lines)
    encoded = urllib.parse.quote(message)
    links = [
        {"number": n, "url": f"https://wa.me/91{n}?text={encoded}"}
        for n in OWNER_WHATSAPP_NUMBERS
    ]
    return {
        "date": target,
        "message": message,
        "owner_numbers": OWNER_WHATSAPP_NUMBERS,
        "links": links,
    }



@api_router.post("/seed")
async def seed_inventory(_: bool = Depends(require_auth)):
    """Seed sample inventory if empty."""
    count = await db.inventory.count_documents({})
    if count > 0:
        return {"seeded": False, "message": "Inventory already has data"}
    samples = [
        ("PENDANT250", "Pendant", "Pendant 250", 250, 100, 80),
        ("PENDANT500", "Pendant", "Pendant 500", 500, 50, 180),
        ("EARRING200", "Earring", "Earring 200", 200, 50, 60),
        ("EARRING400", "Earring", "Earring 400", 400, 30, 150),
        ("RING300", "Ring", "Ring 300", 300, 40, 100),
        ("RING600", "Ring", "Ring 600", 600, 20, 220),
        ("BANGLE800", "Bangle", "Bangle 800", 800, 15, 320),
        ("NECKLACE1200", "Necklace", "Necklace 1200", 1200, 10, 500),
        ("BRACELET450", "Bracelet", "Bracelet 450", 450, 25, 170),
        ("ANKLET350", "Anklet", "Anklet 350", 350, 5, 130),
    ]
    now = now_ist().isoformat()
    for item_id, category, name, price, qty, cost in samples:
        await db.inventory.insert_one({
            "id": str(uuid.uuid4()),
            "item_id": item_id,
            "category": category,
            "item_name": name,
            "price": float(price),
            "cost_price": float(cost),
            "opening_qty": qty,
            "current_qty": qty,
            "sold_qty": 0,
            "created_date": now,
            "last_updated": now,
        })
    return {"seeded": True, "count": len(samples)}


@api_router.get("/")
async def root():
    return {"message": "Iminationz POS API", "store": STORE_NAME}


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
